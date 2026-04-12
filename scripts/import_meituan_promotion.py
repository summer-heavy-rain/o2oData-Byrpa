"""
导入美团推广费（推广美团自营.xlsx → ods_rpa_meituan_promotion）
支持单文件导入：python -m scripts.import_meituan_promotion <路径> [dt]
或直接处理当前目录下 推广美团自营_YYYY-MM-DD.xlsx
"""
import asyncio, ssl, sys, os, glob, datetime
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
load_dotenv(Path(__file__).parent.parent / ".env")
DATABASE_URL = os.environ["DATABASE_URL"]

CREATE_TABLE = """
CREATE TABLE IF NOT EXISTS ods_rpa_meituan_promotion (
    dt            DATE        NOT NULL,
    store_id      BIGINT,
    store_name    TEXT,
    store_type    TEXT,
    amount        NUMERIC(10,2),  -- 正数，推广消费金额
    _source_file  TEXT,
    _load_time    TIMESTAMPTZ DEFAULT NOW(),
    PRIMARY KEY (dt, store_id)
);
"""


def read_promotion_xlsx(path: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True)

    # 门店ID → 简称映射
    store_map = {}
    ws_store = wb["门店ID"]
    for i, row in enumerate(ws_store.iter_rows(values_only=True)):
        if i == 0:
            continue
        short_name, meituan_id = row[1], row[10]
        if meituan_id:
            store_map[int(meituan_id)] = short_name

    ws = wb["推广费流水"]
    rows = []
    source_file = Path(path).name
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        store_id = row[0]
        date_str = row[8]   # 转换日期（业务归属日期，跨日扣款按此归日，规则 #39）
        amount_raw = row[9]  # 列1 = VALUE(金额)，负数
        store_name = row[7]  # 店铺（XLOOKUP 结果）

        if not store_id or not date_str:
            continue

        # 解析日期
        try:
            dt = datetime.datetime.strptime(str(date_str), "%Y-%m-%d").date()
        except ValueError:
            # 尝试 "YYYY-M-DD" 格式
            parts = str(date_str).split("-")
            dt = datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))

        amount = abs(float(amount_raw)) if amount_raw else 0.0
        rows.append((dt, int(store_id), store_name or store_map.get(int(store_id), "未匹配"), "直营", amount, source_file))
    return rows


async def main(xlsx_path: str):
    import asyncpg

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    rows = read_promotion_xlsx(xlsx_path)
    print(f"读取到 {len(rows)} 行推广费数据")
    if rows:
        print(f"  日期：{rows[0][0]}，门店示例：{rows[0][2]}，金额：{rows[0][4]}")

    conn = await asyncpg.connect(DATABASE_URL, ssl=ctx)
    try:
        await conn.execute(CREATE_TABLE)
        # 用 ON CONFLICT 做 upsert（同日同门店覆盖）
        await conn.executemany(
            """
            INSERT INTO ods_rpa_meituan_promotion
                (dt, store_id, store_name, store_type, amount, _source_file)
            VALUES ($1, $2, $3, $4, $5, $6)
            ON CONFLICT (dt, store_id) DO UPDATE SET
                store_name   = EXCLUDED.store_name,
                amount       = EXCLUDED.amount,
                _source_file = EXCLUDED._source_file,
                _load_time   = NOW()
            """,
            rows,
        )
        count = await conn.fetchval(
            "SELECT COUNT(*) FROM ods_rpa_meituan_promotion WHERE dt = $1",
            rows[0][0]
        )
        total = await conn.fetchval("SELECT SUM(amount) FROM ods_rpa_meituan_promotion WHERE dt = $1", rows[0][0])
        print(f"✅ ods_rpa_meituan_promotion 写入完成：{count} 条，合计推广费 ¥{total:.2f}")
    finally:
        await conn.close()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        # 自动找当前目录下最新的推广文件
        files = glob.glob(str(Path(__file__).parent.parent / "推广美团自营_*.xlsx"))
        if not files:
            print("❌ 未找到推广美团自营_*.xlsx 文件")
            sys.exit(1)
        path = sorted(files)[-1]
        print(f"自动选择文件: {path}")

    asyncio.run(main(path))
