"""生成字段校对 Excel（多 sheet），给 RPA 组和阿潘用"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path
import pandas as pd

YELLOW = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
HEADER_FILL = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
BODY_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
RED_FONT = Font(bold=True, size=10, color='CC0000')
RED_FILL = PatternFill(start_color='FFFF6666', end_color='FFFF6666', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

NAS = r"\\192.168.1.49\数仓rpa文件\数据组文件"
DT = "2026-03-25"


def add_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN_BORDER


def make_sheet(ws, ods_table, rpa_file, pattern, all_cols, needed_set, notes=''):
    """
    all_cols: list of column names from actual file
    needed_set: set of column names that are needed (yellow)
    """
    headers = ['序号', '原始列名(来自3.25实际文件)', '是否建模需要', '阿潘校对(打✓或✗)', '备注']
    add_header(ws, headers)

    for i, col_name in enumerate(all_cols, 1):
        row = i + 1
        is_needed = col_name in needed_set
        vals = [i, col_name, '✓ 需要' if is_needed else '', '', '']
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if is_needed and col_idx <= 3:
                c.fill = YELLOW

    last = len(all_cols) + 3
    ws.cell(row=last, column=1, value='ODS目标表:').font = BOLD_FONT
    ws.cell(row=last, column=2, value=ods_table).font = BODY_FONT
    ws.cell(row=last+1, column=1, value='3.25 RPA文件:').font = BOLD_FONT
    ws.cell(row=last+1, column=2, value=rpa_file).font = BODY_FONT
    ws.cell(row=last+2, column=1, value='规范文件名:').font = BOLD_FONT
    ws.cell(row=last+2, column=2, value=pattern).font = BODY_FONT
    if notes:
        ws.cell(row=last+3, column=1, value='注意:').font = BOLD_FONT
        ws.cell(row=last+3, column=2, value=notes).font = RED_FONT

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30


def read_openpyxl_headers(path, sheet=0, max_row=1):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if isinstance(sheet, str):
        ws = wb[sheet]
    else:
        ws = wb.worksheets[sheet]
    cols = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=max_row, column=col).value
        cols.append(str(v).strip() if v else f'col_{col}')
    wb.close()
    return cols


def read_multi_header(path, sheet, num_rows=2):
    """Read multi-level headers, join with →"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    cols = []
    for col in range(1, ws.max_column + 1):
        parts = []
        for row in range(1, num_rows + 1):
            v = ws.cell(row=row, column=col).value
            if v and str(v).strip():
                parts.append(str(v).strip())
        deduped = []
        for p in parts:
            if not deduped or deduped[-1] != p:
                deduped.append(p)
        cols.append(' → '.join(deduped) if deduped else f'col_{col}')
    wb.close()
    return cols


wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

# ============================================================
# 1. 麦芽田
# ============================================================
print('1. 麦芽田订单...')
try:
    df = pd.read_excel(f'{NAS}/【数据组】麦芽田数据导出/{DT}/麦芽田-报表中心（小猴）.xlsx', nrows=0)
    myt_cols = list(df.columns)
except:
    myt_cols = ['流水号','来源平台','平台店铺','状态','原流水号','订单编号','是否预约',
        '下单日期','期望送达','预计发货时间','完成时间','备注','收货人','收货人电话',
        '地址','订单总金额','商家实收金额','配送门店','距离','配送平台','配送单号',
        '骑手姓名','骑手电话','配送费','小费','总配送费','配送距离','配送状态','实际发货时间']

myt_needed = {'来源平台','平台店铺','状态','订单编号','下单日期','距离','总配送费'}
ws = wb_out.create_sheet('麦芽田订单')
make_sheet(ws, 'ods_rpa_maiyatian_delivery',
    '麦芽田-报表中心（小猴）.xlsx', '麦芽田-报表中心（{门店名}）.xlsx',
    myt_cols, myt_needed)
print(f'   {len(myt_cols)} 列, {len(myt_needed)} 需要')

# ============================================================
# 2. 美团账单
# ============================================================
print('2. 美团账单...')
mt_cols = read_multi_header(
    f'{NAS}/【数据组】美团数据导出/{DT}/美团-账单（小猴）.xlsx', '订单明细', 2)
mt_needed = set()
needed_kw = ['门店id','门店名称','物理城市','交易类型','交易描述','订单号','下单时间',
    '订单状态','商家应收款','商品总价','用户支付配送费','餐盒费','打包袋',
    '商家活动总支出','公益捐款','佣金','佣金2','健康卡费用商家部分',
    '配送方式','配送服务费','配送费返利']
for c in mt_cols:
    for kw in needed_kw:
        if kw == c or c.endswith(kw) or c.endswith(f'→ {kw}') or f' → {kw}' in c:
            mt_needed.add(c)
            break

ws = wb_out.create_sheet('美团账单')
make_sheet(ws, 'ods_rpa_meituan_fin_order',
    '美团-账单（小猴）.xlsx → 订单明细', '美团-账单（{门店名}）.xlsx',
    mt_cols, mt_needed, '多级表头(2行)')
print(f'   {len(mt_cols)} 列, {len(mt_needed)} 需要')

# ============================================================
# 3. 美团商品明细
# ============================================================
print('3. 美团商品明细...')
mt_prod_cols = list(pd.read_csv(
    f'{NAS}/【数据组】美团数据导出/{DT}/美团-商品明细（小猴）.csv',
    encoding='gbk', nrows=0).columns)
mt_prod_needed = {'订单编号','下单时间','商家名称','商家ID','商家所在城市',
    '订单状态','UPC码','店内码/货号','商品销售数量','部分退款商品数量'}
ws = wb_out.create_sheet('美团商品明细')
make_sheet(ws, 'ods_rpa_meituan_product_detail',
    '美团-商品明细（小猴）.csv', '美团-商品明细（{门店名}）.csv',
    mt_prod_cols, mt_prod_needed, 'CSV GBK编码')
print(f'   {len(mt_prod_cols)} 列, {len(mt_prod_needed)} 需要')

# ============================================================
# 4. 饿了么账单 (4级表头)
# ============================================================
print('4. 饿了么账单...')
elm_cols = read_multi_header(
    f'{NAS}/【数据组】饿了么数据导出/{DT}/饿了么-财务账单.xlsx', '销售账单明细', 4)
elm_needed_pos = {1,2,3,7,8,10,11,12,14,18,19,20,27,28,56,60,63,66,69,75,80,84}
elm_needed = {elm_cols[p-1] for p in elm_needed_pos if p <= len(elm_cols)}
ws = wb_out.create_sheet('饿了么账单')
make_sheet(ws, 'ods_rpa_eleme_fin_sales_detail',
    '饿了么-财务账单.xlsx → 销售账单明细', '饿了么-财务账单{(账号)}.xlsx',
    elm_cols, elm_needed, '4级表头! 有重名列(实收服务费×2)')
print(f'   {len(elm_cols)} 列, {len(elm_needed)} 需要')

# ============================================================
# 5. 饿了么订单
# ============================================================
print('5. 饿了么订单...')
elo_cols = read_openpyxl_headers(
    f'{NAS}/【数据组】饿了么数据导出/{DT}/饿了么-订单.xlsx', '订单导出')
elo_needed = set()
elo_kw = ['订单来源','订单编号','城市','商户名称','门店ID',
    '订单状态','退款状态','下单时间','商品自定义ID','商品条形码','购买数量']
for c in elo_cols:
    if c in elo_kw:
        elo_needed.add(c)
ws = wb_out.create_sheet('饿了么订单')
make_sheet(ws, 'ods_rpa_eleme_order',
    '饿了么-订单.xlsx → 订单导出', '饿了么-订单{(账号)}.xlsx',
    elo_cols, elo_needed)
print(f'   {len(elo_cols)} 列, {len(elo_needed)} 需要')

# ============================================================
# 6. 京东小时达财务
# ============================================================
print('6. 京东小时达...')
jd_fin_cols = list(pd.read_excel(
    f'{NAS}/【数据组】京东数据导出/{DT}/账单下载.xlsx',
    sheet_name='sku对账单下载', nrows=0).columns)
jd_fin_needed = {'到家业务单号','业务类型','门店编号','费用类型','结算金额','下单时间'}
ws = wb_out.create_sheet('京东小时达财务')
make_sheet(ws, 'ods_rpa_jd_finance',
    '账单下载.xlsx → sku对账单下载', '账单下载.xlsx',
    jd_fin_cols, jd_fin_needed)
print(f'   {len(jd_fin_cols)} 列, {len(jd_fin_needed)} 需要')

# ============================================================
# 7. 京东订单
# ============================================================
print('7. 京东订单...')
jd_order_dfs = pd.read_html(
    f'{NAS}/【数据组】京东数据导出/{DT}/订单查询.xlsx', encoding='utf-8')
jd_df = jd_order_dfs[0]
first_col_name = str(jd_df.columns[0]).strip()
if first_col_name.isdigit():
    jd_order_cols = [str(v).strip() if pd.notna(v) else f'col_{i}' for i, v in enumerate(jd_df.iloc[0].values)]
else:
    jd_order_cols = [str(c) for c in jd_df.columns]
jd_order_needed = {'订单编号','门店ID','门店名称','商家sku','UPC','商品数量','接单时间','订单来源'}
ws = wb_out.create_sheet('京东订单')
make_sheet(ws, 'ods_rpa_jd_order',
    '订单查询.xlsx', '订单查询.xlsx',
    jd_order_cols, jd_order_needed, '实际是HTML伪装成xlsx')
print(f'   {len(jd_order_cols)} 列, {len(jd_order_needed)} 需要')

# ============================================================
# 8. 美团推广
# ============================================================
print('8. 美团推广...')
promo_wb = openpyxl.load_workbook(
    f'{NAS}/【数据组】美团数据导出/{DT}/推广美团自营.xlsx', read_only=True, data_only=True)
for sname in ['余额流水', '推广费流水']:
    pws = promo_wb[sname]
    p_cols = []
    for col in range(1, pws.max_column + 1):
        v = pws.cell(row=1, column=col).value
        p_cols.append(str(v).strip() if v else f'col_{col}')
    p_needed = set(p_cols)
    ws = wb_out.create_sheet(f'美团推广-{sname}')
    make_sheet(ws, 'ods_rpa_meituan_promotion',
        f'推广美团自营.xlsx → {sname}', '美团-推广账单（直营/加盟）.xlsx',
        p_cols, p_needed, '全部字段都要,不裁剪')
    print(f'   推广-{sname}: {len(p_cols)} 列(全要)')
promo_wb.close()

# ============================================================
# 9. RPA文件名规范
# ============================================================
print('9. RPA文件名规范...')
ws = wb_out.create_sheet('⚠RPA文件名规范')
rpa_headers = ['平台', '文件类型', '规范文件名(请严格遵守)', '3月25日实际文件名', '近期实际文件名', '问题']
add_header(ws, rpa_headers)
rpa_data = [
    ['美团', '商品明细', '美团-商品明细（{门店名}）.csv', '美团-商品明细（小猴）.csv',
     '小猴快跑医疗器械美团-订单.csv(4.7)', '文件名完全变了,入库失败!'],
    ['美团', '账单', '美团-账单（{门店名}）.xlsx', '美团-账单（小猴）.xlsx',
     '小猴快跑医疗器械美团-账单.xlsx(4.7)', '文件名变了,入库失败!'],
    ['美团', '推广(直营)', '美团-推广账单（直营）.xlsx', '推广美团自营.xlsx',
     '美团直营店推广账单.xlsx(4.8)', '每次名字都不一样!'],
    ['美团', '推广(加盟)', '美团-推广账单（加盟）.xlsx', '推广美团加盟.xlsx',
     '美团加盟店推广账单.xlsx(4.8)', '每次名字都不一样!'],
    ['美团', '⚠门店拆分', '华信和槐荫必须分别导出,不要合并', '华信/槐荫各一个文件',
     '合并成"善培臣"(4.6-4.7)', '合并导致门店维度丢失!'],
    ['饿了么', '订单', '饿了么-订单{(账号)}.xlsx', '饿了么-订单.xlsx', '待确认', ''],
    ['饿了么', '推广', '饿了么-推广数据{(账号)}.xlsx', '饿了么-推广数据.xlsx', '待确认', ''],
    ['饿了么', '财务', '饿了么-财务账单{(账号)}.xlsx', '饿了么-财务账单.xlsx', '待确认', ''],
    ['京东', '订单', '订单查询.xlsx', '订单查询.xlsx', '待确认', ''],
    ['京东', 'SKU对账单', '账单下载.xlsx', '账单下载.xlsx', '待确认', ''],
    ['麦芽田', '配送报表', '麦芽田-报表中心（{门店名}）.xlsx',
     '麦芽田-报表中心（小猴）.xlsx', '待确认', ''],
]
for i, row_data in enumerate(rpa_data, 2):
    for col_idx, v in enumerate(row_data, 1):
        c = ws.cell(row=i, column=col_idx, value=v)
        c.font = BODY_FONT
        c.border = THIN_BORDER
        if col_idx == 6 and v:
            c.fill = RED_FILL
            c.font = RED_FONT

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 42
ws.column_dimensions['D'].width = 34
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 32

note_row = len(rpa_data) + 3
ws.cell(row=note_row, column=1, value='⚠ RPA组注意:').font = Font(bold=True, size=13, color='CC0000')
ws.cell(row=note_row+1, column=1,
    value='1. 文件名必须严格按「规范文件名」列的格式导出，花括号部分替换为实际值，其余一字不改').font = BOLD_FONT
ws.cell(row=note_row+2, column=1,
    value='2. 不要合并门店文件！华信和槐荫必须分开导出，不要合并成"善培臣"').font = RED_FONT
ws.cell(row=note_row+3, column=1,
    value='3. 不要在文件名里加日期、长串ID、公司全称等额外信息').font = BOLD_FONT

# ============================================================
# Save
# ============================================================
out_path = Path(__file__).parent.parent / 'docs' / '字段校对表_给RPA组和阿潘_v3.xlsx'
wb_out.save(str(out_path))
print(f'\n✅ 已保存: {out_path}')
print(f'共 {len(wb_out.sheetnames)} 个sheet: {wb_out.sheetnames}')
