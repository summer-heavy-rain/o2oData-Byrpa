#!/usr/bin/env python3
"""京东到家「商家对账单」宽表：应结金额恒等式校验。

三平台两层校验铁律（见 docs/京东到家商家对账单验数.md §1.1）：
  第一层：订单号匹配 — 订单明细中有但财务账单无的订单 → 暂存不计入当期，
         单独列出（按门店汇总）供 xiaohou 核查，待账单更新后补算。
  第二层：公式推导 — 已匹配订单的各列代数和 vs 应结金额逐单比对，
         不等 → 新增费用类型未覆盖，必须通知 xiaohou。

本脚本针对单份 xlsx 执行第二层校验（xlsx 本身即为财务账单，不涉及订单匹配）。
第一层匹配在数仓 SQL / 批量对账脚本中实现。

用法:
  python scripts/validate_jd_daojia_merchant_statement.py /path/to/商家对账单_xxx.xlsx

依赖: openpyxl
"""

from __future__ import annotations

import argparse
import sys
from decimal import Decimal
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 与 docs/京东到家商家对账单验数.md 一致
# 开票金额：永不纳入应结推导（勿加入下列任一组）
COLS_BASE = [
    "订单原价",
    "平台承担补贴(市场费)",
    "商家承担货款补贴",
    "商家承担运费补贴",
    "商家自送配送费",
    "商家承担小费",
    "取件服务费(开票)(正向单展示远距离运费;售后单则展示达达售后运费)",
    "总佣金(货款佣金+运费佣金+餐盒费佣金)(可开票)",
    "基础服务费",
]
# 表头存在时才参与本期求和（见文档「账单未出现」规则）；均为「到家」宽表列
COLS_OPTIONAL = (
    "阶梯扣点佣金",  # 并入平台扣点佣金
    "包装费",  # 仅到家宽表；小时达包装费为长表费用类型单独分项，见文档
)


def _num(x: object) -> Decimal:
    if x is None or x == "":
        return Decimal(0)
    if isinstance(x, (int, float)):
        return Decimal(str(x))
    s = str(x).strip().replace(",", "")
    if s == "" or s == "--":
        return Decimal(0)
    return Decimal(s)


def main() -> int:
    ap = argparse.ArgumentParser(description="校验京东到家商家对账单应结恒等式")
    ap.add_argument("xlsx", type=Path, help="商家对账单 xlsx 路径")
    ap.add_argument(
        "--sheet",
        default="第1页",
        help="明细表 sheet 名（默认 第1页）",
    )
    args = ap.parse_args()
    path = args.xlsx.expanduser()
    if not path.is_file():
        print(f"文件不存在: {path}", file=sys.stderr)
        return 2

    wb = load_workbook(path, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"无 sheet: {args.sheet}，现有: {wb.sheetnames}", file=sys.stderr)
        wb.close()
        return 2
    ws = wb[args.sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("空表", file=sys.stderr)
        wb.close()
        return 2
    hdr = list(rows[0])
    missing = [c for c in COLS_BASE + ["应结金额", "订单号"] if c not in hdr]
    if missing:
        print("缺少列:", missing, file=sys.stderr)
        wb.close()
        return 2

    cols_sum = list(COLS_BASE)
    skipped_optional: list[str] = []
    for opt in COLS_OPTIONAL:
        if opt in hdr:
            cols_sum.append(opt)
        else:
            skipped_optional.append(opt)
            print(
                f"提示: 表头无「{opt}」，本期不参与求和",
                file=sys.stderr,
            )

    idx = {h: hdr.index(h) for h in cols_sum + ["应结金额", "订单号"]}
    tol = Decimal("0.02")
    fails: list[tuple[object, Decimal, Decimal]] = []
    sum_yj = Decimal(0)

    for rno, row in enumerate(rows[1:], start=2):
        if not row or row[idx["订单号"]] is None:
            continue
        oid = row[idx["订单号"]]
        parts = sum(_num(row[idx[c]]) for c in cols_sum)
        yj = _num(row[idx["应结金额"]])
        sum_yj += yj
        if abs(yj - parts) > tol:
            fails.append((oid, yj, parts))

    # ── 三平台铁律：推导 ≠ 账单 → 必须通知 xiaohou ──
    if fails:
        print(f"\n⚠️  明细行校验: 失败 {len(fails)} 笔 — 公式推导 ≠ 应结金额")
        print("【三平台铁律】差异说明存在新增费用类型/列未被公式覆盖，")
        print("必须通知 xiaohou 确认后才能更新公式，禁止静默跳过。")
        if skipped_optional:
            print(f"本期跳过的可选列: {skipped_optional}（若差异与其相关，需确认是否应纳入）")
        for oid, yj, parts in fails:
            print(f"  订单号={oid} 应结={yj} 推导={parts} diff={yj - parts}")
    else:
        print(f"\n✅ 明细行校验: 全部通过 ({sum(1 for r in rows[1:] if r and r[idx['订单号']] is not None)} 笔)")

    if "第2页" in wb.sheetnames:
        ws2 = wb["第2页"]
        r2 = list(ws2.iter_rows(values_only=True))
        if len(r2) >= 2 and r2[0] and r2[0][0] == "合计应结金额":
            total = _num(r2[1][0])
            print(f"第1页 SUM(应结金额)={sum_yj}  第2页合计应结金额={total}  一致={abs(sum_yj - total) <= tol}")
        else:
            print("第2页格式非预期，跳过合计交叉校验")

    wb.close()
    return 1 if fails else 0


if __name__ == "__main__":
    raise SystemExit(main())
