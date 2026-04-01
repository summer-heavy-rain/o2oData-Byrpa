"""
多格式文件统一读取器
支持: xlsx / csv(多编码) / html-as-xlsx / xml-as-xlsx / 多sheet / 合并表头
返回: list[DataFrame]，每个 DataFrame 对应一个 ODS 目标表
"""
from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from io import StringIO
from pathlib import Path

import pandas as pd

from .config_loader import FileRule, SheetRule


def read_file(
    path: Path,
    rule: FileRule,
) -> list[tuple[str, pd.DataFrame]]:
    """
    读取单个文件，返回 [(target_table, DataFrame), ...]
    多 sheet 文件返回多个元组
    """
    fmt = rule.format

    if fmt == "xlsx":
        return _read_xlsx_single(path, rule)
    elif fmt == "xlsx_multi_sheet":
        return _read_xlsx_multi(path, rule)
    elif fmt == "csv":
        return _read_csv(path, rule)
    elif fmt == "html_as_xlsx":
        return _read_html_as_xlsx(path, rule)
    elif fmt == "xml_as_xlsx":
        return _read_xml_as_xlsx(path, rule)
    else:
        raise ValueError(f"未知文件格式: {fmt}")


def _read_xlsx_single(path: Path, rule: FileRule) -> list[tuple[str, pd.DataFrame]]:
    sheet = rule.sheet if rule.sheet is not None else 0
    df = pd.read_excel(
        path,
        sheet_name=sheet,
        header=rule.header_row - 1,
        dtype=str,
        engine="openpyxl",
    )
    raw_count = len(df.columns)
    df.columns = _sanitize_columns(df.columns.tolist())
    attach_col_count(df, raw_count)
    return [(rule.target_table, df)]


def _read_xlsx_multi(path: Path, rule: FileRule) -> list[tuple[str, pd.DataFrame]]:
    results = []
    for sheet_rule in rule.sheets:
        try:
            if sheet_rule.header_rows > 1:
                df = _read_merged_header_sheet(path, sheet_rule)
            else:
                skip = [r - 1 for r in sheet_rule.skip_rows]
                header_idx = sheet_rule.header_row - 1
                skipped_before = sum(1 for s in skip if s < header_idx)
                adjusted_header = header_idx - skipped_before
                df = pd.read_excel(
                    path,
                    sheet_name=sheet_rule.sheet,
                    header=adjusted_header,
                    skiprows=skip if skip else None,
                    dtype=str,
                    engine="openpyxl",
                )
            raw_count = len(df.columns)
            df.columns = _sanitize_columns(df.columns.tolist())
            attach_col_count(df, raw_count)
            df = df.dropna(how="all")
            if not df.empty:
                results.append((sheet_rule.target_table, df))
        except (ValueError, KeyError):
            pass
    return results


def _read_merged_header_sheet(path: Path, rule: SheetRule) -> pd.DataFrame:
    """处理合并表头（两行表头展平为 group_sub 格式）
    自动解析合并单元格：被 merge 遮蔽的表头文本会恢复为正确值，
    避免不同文件列数不同时 col_{i} 绝对位置错位。
    """
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb[rule.sheet] if isinstance(rule.sheet, str) else wb.worksheets[rule.sheet]

    merged_lookup = _build_merged_lookup(ws)

    hr = rule.header_row  # 1-based
    max_col = ws.max_column
    row1_raw = [_resolve_cell(ws, merged_lookup, hr, c + 1) for c in range(max_col)]
    row2_raw = [_resolve_cell(ws, merged_lookup, hr + 1, c + 1) for c in range(max_col)]
    wb.close()

    group = ""
    columns = []
    for i, (g, s) in enumerate(zip(row1_raw, row2_raw)):
        if g is not None:
            group = str(g).strip()
        sub = str(s).strip() if s is not None else f"col_{i}"
        if rule.merge_strategy == "group_sub" and group and group != sub:
            col_name = f"{group}_{sub}"
        else:
            col_name = sub
        columns.append(col_name)

    data_start = (hr - 1) + rule.header_rows
    df = pd.read_excel(
        path,
        sheet_name=rule.sheet,
        header=None,
        skiprows=list(range(data_start)),
        dtype=str,
        engine="openpyxl",
    )
    if len(df.columns) <= len(columns):
        df.columns = columns[: len(df.columns)]
    else:
        extra = [f"_extra_{i}" for i in range(len(df.columns) - len(columns))]
        df.columns = columns + extra

    df.columns = _sanitize_columns(df.columns.tolist())
    attach_col_count(df, max_col)
    return df.dropna(how="all")


def _build_merged_lookup(ws) -> dict[tuple[int, int], object]:
    """Pre-build (row, col) -> primary_cell_value for all merged ranges."""
    lookup: dict[tuple[int, int], object] = {}
    for mr in ws.merged_cells.ranges:
        primary_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    lookup[(r, c)] = primary_val
    return lookup


def _resolve_cell(ws, merged_lookup: dict, row: int, col: int):
    """Get effective cell value, resolving merged cells to their primary value."""
    val = ws.cell(row, col).value
    if val is not None:
        return val
    return merged_lookup.get((row, col))


def _read_csv(path: Path, rule: FileRule) -> list[tuple[str, pd.DataFrame]]:
    for enc in [rule.encoding, "gbk", "gb18030", "utf-8-sig", "utf-8", "latin1"]:
        try:
            df = pd.read_csv(
                path,
                encoding=enc,
                header=rule.header_row - 1,
                dtype=str,
            )
            raw_count = len(df.columns)
            df.columns = _sanitize_columns(df.columns.tolist())
            attach_col_count(df, raw_count)
            return [(rule.target_table, df)]
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f"无法以任何编码读取 CSV: {path}")


def _read_html_as_xlsx(path: Path, rule: FileRule) -> list[tuple[str, pd.DataFrame]]:
    """京东等平台导出的 xlsx 实际是 HTML table"""
    with open(path, "r", encoding=rule.encoding, errors="replace") as f:
        html = f.read()

    tables = pd.read_html(StringIO(html), header=0)
    if not tables:
        raise ValueError(f"HTML 文件中未找到表格: {path}")

    df = tables[0].astype(str).replace("nan", pd.NA)
    raw_count = len(df.columns)
    df.columns = _sanitize_columns(df.columns.tolist())
    attach_col_count(df, raw_count)
    return [(rule.target_table, df)]


def _read_xml_as_xlsx(path: Path, rule: FileRule) -> list[tuple[str, pd.DataFrame]]:
    """麦芽田等导出的 xlsx 实际是 XML SpreadsheetML"""
    with open(path, "r", encoding=rule.encoding, errors="replace") as f:
        content = f.read()

    content = re.sub(r"&(?!amp;|lt;|gt;|quot;|apos;|#)\w+;", "", content)

    root = ET.fromstring(content)

    worksheets = root.findall(".//{urn:schemas-microsoft-com:office:spreadsheet}Worksheet")
    if not worksheets:
        raise ValueError(f"XML 中未找到 Worksheet: {path}")

    ws = worksheets[0]
    rows_el = ws.findall(".//{urn:schemas-microsoft-com:office:spreadsheet}Row")

    all_rows = []
    for row_el in rows_el:
        cells = row_el.findall("{urn:schemas-microsoft-com:office:spreadsheet}Cell")
        row_data = []
        col_idx = 0
        for cell in cells:
            idx_attr = cell.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
            if idx_attr:
                target_idx = int(idx_attr) - 1
                while col_idx < target_idx:
                    row_data.append(None)
                    col_idx += 1
            data_el = cell.find("{urn:schemas-microsoft-com:office:spreadsheet}Data")
            row_data.append(data_el.text if data_el is not None else None)
            col_idx += 1
        all_rows.append(row_data)

    if not all_rows:
        raise ValueError(f"XML 文件无数据: {path}")

    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(all_rows[0])]
    max_cols = max(len(r) for r in all_rows)
    while len(headers) < max_cols:
        headers.append(f"col_{len(headers)}")

    data = []
    for row in all_rows[1:]:
        padded = row + [None] * (max_cols - len(row))
        data.append(padded[:max_cols])

    df = pd.DataFrame(data, columns=headers, dtype=str)
    raw_count = len(headers)
    df.columns = _sanitize_columns(df.columns.tolist())
    attach_col_count(df, raw_count)
    return [(rule.target_table, df.dropna(how="all"))]


def attach_col_count(df: pd.DataFrame, raw_col_count: int) -> pd.DataFrame:
    """将源文件原始列数附加到 DataFrame 属性上（不加列，由 ingest 决定是否写入）"""
    df.attrs["_raw_col_count"] = raw_col_count
    return df


def _sanitize_columns(columns: list) -> list[str]:
    """列名清洗：去空白、去括号特殊字符、去重"""
    seen: dict[str, int] = {}
    result = []
    for col in columns:
        col = str(col).strip()
        col = col.replace("（", "_").replace("）", "")
        col = col.replace("(", "_").replace(")", "")
        col = col.replace("【", "").replace("】", "")
        col = col.replace("/", "_")
        col = re.sub(r"\s+", "_", col)
        col = re.sub(r"_+", "_", col).strip("_")
        if not col:
            col = "unnamed"
        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
        result.append(col)
    return result
